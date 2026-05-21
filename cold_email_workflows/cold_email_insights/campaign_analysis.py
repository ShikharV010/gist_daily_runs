"""
campaign_analysis.py
────────────────────────────────────────────────────────────────────
Produces data/campaign_analysis.xlsx — a formula-driven side-by-side of
the 10 industries the team is evaluating against the Manufacturing
benchmark.

Architecture (so the workbook stays editable):

  Sheet 1 : Variables
            One row per Sequencer campaign with every raw count the
            dashboard surfaces (emails, leads, replies, bounces,
            interested, demos, shows, pending, no-shows, completed,
            closes, ARR, MRR). Plus an Industry Group column so Google
            Ads (Running) + Google Ads (Stopped) roll up together.

  Sheet 2 : Readme
            How to use the workbook + glossary of metric definitions.

  Sheets 3+ : One per industry group. Each pulls totals from the
            Variables sheet via SUMIFS — no hardcoded numbers. Layout:
              Metric | <Industry> | Manufacturing (benchmark) | Index vs MFG

If you edit a number on Variables, every industry sheet updates on
recalc. Run again whenever metrics.json changes to refresh the raw inputs.
"""
import json, os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
METRICS = os.path.join(ROOT, 'data', 'metrics.json')
OUTFILE = os.path.join(ROOT, 'data', 'campaign_analysis.xlsx')

TODAY = datetime.now().strftime('%Y-%m-%d')

# Industries the user wants pages for. Each entry is (page-title, [industry-names-in-data]).
# Most groups map 1:1 to the industry tag the ETL writes. Google Ads collapses two.
INDUSTRY_GROUPS = [
    ('Business Services (1-5)',  ['Business Services (1-5)']),
    ('Mix Services (1-5)',       ['Mix Services (1-5)']),
    ('Mix Services (5-200)',     ['Mix Services (5-200)']),
    ('MFG SE Hook',              ['MFG SE Hook']),
    ('MFG Account',              ['MFG Account']),
    ('Corporate Training',       ['Corporate Training']),
    ('MFG 1-5',                  ['MFG 1-5']),
    ('Financial Services 1-5',   ['Financial Services 1-5']),
    ('Construction',             ['Construction']),
    ('Google Ads',               ['Google Ads (Running)', 'Google Ads (Stopped)']),
]
BENCHMARK = 'Manufacturing'

# ── Load metrics.json ─────────────────────────────────────────────────────
m = json.load(open(METRICS))
print(f'metrics.json: generated {m["generated_at"]}', flush=True)

# Per-campaign rollups from demo_bookings (the campaigns array itself only
# carries email/lead/replied/bounced/interested totals — demo & close counts
# need to be aggregated from demo_bookings filtered by campaign_id).
demos     = defaultdict(int)
showups   = defaultdict(int)
pending   = defaultdict(int)
completed = defaultdict(int)
closed    = defaultdict(int)
arr       = defaultdict(float)
mrr       = defaultdict(float)

for b in m['demo_bookings']:
    cid = b.get('campaign_id')
    if not cid:
        continue
    demos[cid] += 1
    ss = b.get('show_status')
    dd = b.get('demo_scheduled_date') or ''
    if ss == 'Y':
        showups[cid] += 1
        completed[cid] += 1
    elif ss in ('N', 'R'):
        completed[cid] += 1
    elif ss == 'P':
        if dd >= TODAY:
            pending[cid] += 1
        else:
            completed[cid] += 1  # past P = treat as completed (no-show)
    if b.get('closed'):
        closed[cid] += 1
        arr[cid] += b.get('arr') or 0
        mrr[cid] += b.get('monthly_amount') or 0

# Build a single industry_group key per campaign matching INDUSTRY_GROUPS
INDUSTRY_TO_GROUP = {}
for group_title, industries in INDUSTRY_GROUPS:
    for ind in industries:
        INDUSTRY_TO_GROUP[ind] = group_title
INDUSTRY_TO_GROUP[BENCHMARK] = BENCHMARK  # benchmark stays as-is

# ── Styling primitives ────────────────────────────────────────────────────
H_FONT   = Font(bold=True, color='FFFFFFFF', size=11)
H_FILL   = PatternFill('solid', fgColor='FF0070FF')
SECT_FILL = PatternFill('solid', fgColor='FFDBEAFE')
SECT_FONT = Font(bold=True, color='FF1E40AF', size=10)
THIN     = Side(style='thin', color='FFD0D5DD')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_L  = Alignment(horizontal='left',  vertical='center')
ALIGN_R  = Alignment(horizontal='right', vertical='center', wrap_text=False)
BOLD     = Font(bold=True)
INDEX_HI = PatternFill('solid', fgColor='FFD1FAE5')  # green wash for >100%
INDEX_LO = PatternFill('solid', fgColor='FFFEE2E2')  # red wash for <50%

def style_header(ws, row=1):
    for c in ws[row]:
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = ALIGN_L
        c.border = BORDER

def autosize(ws, max_w=42):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_w)

# ── Sheet: Variables ──────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

VAR = wb.create_sheet('Variables')
VAR_COLS = [
    'campaign_id', 'campaign_name', 'industry', 'industry_group', 'status',
    'emails_sent', 'leads_contacted', 'replied', 'bounced', 'interested',
    'demos_booked', 'showups', 'pending', 'noshow', 'completed_demos',
    'closed', 'arr', 'mrr',
]
VAR.append(VAR_COLS)
style_header(VAR)

# All campaigns the user wants reflected (the 10 groups + Manufacturing benchmark)
included_industries = set([BENCHMARK])
for _, inds in INDUSTRY_GROUPS:
    included_industries.update(inds)

rows_written = 0
for c in m['campaigns']:
    ind = c.get('industry') or ''
    if ind not in included_industries:
        continue
    cid       = c['id']
    n_demos   = demos.get(cid, 0)
    n_show    = showups.get(cid, 0)
    n_done    = completed.get(cid, 0)
    n_pend    = pending.get(cid, 0)
    n_noshow  = max(0, n_done - n_show)
    VAR.append([
        cid, c['name'], ind, INDUSTRY_TO_GROUP.get(ind, ''), c.get('status') or '',
        c.get('emails_sent') or 0,
        c.get('total_leads_contacted') or 0,
        c.get('replied') or 0,
        c.get('bounced') or 0,
        c.get('interested') or 0,
        n_demos, n_show, n_pend, n_noshow, n_done,
        closed.get(cid, 0), arr.get(cid, 0.0), mrr.get(cid, 0.0),
    ])
    rows_written += 1

print(f'  Variables: {rows_written} campaigns', flush=True)
VAR.freeze_panes = 'A2'
autosize(VAR, max_w=46)

# Column letter map for SUMIFS references
COL = {name: get_column_letter(i + 1) for i, name in enumerate(VAR_COLS)}
LAST_ROW = rows_written + 1  # +1 for header

def sumif(target_col, group_value_cell):
    """SUMIFS on the Variables sheet aggregating `target_col` where
    industry_group == group_value_cell on this sheet."""
    return (
        f'=SUMIFS(Variables!{COL[target_col]}2:{COL[target_col]}{LAST_ROW},'
        f' Variables!{COL["industry_group"]}2:{COL["industry_group"]}{LAST_ROW},{group_value_cell})'
    )

def sumif_literal(target_col, group_literal):
    return (
        f'=SUMIFS(Variables!{COL[target_col]}2:{COL[target_col]}{LAST_ROW},'
        f' Variables!{COL["industry_group"]}2:{COL["industry_group"]}{LAST_ROW},"{group_literal}")'
    )

# ── Helper: build an industry page ────────────────────────────────────────
# Each metric row is (label, formula-builder, format-code, section).
# formula-builder takes (industry_cell_ref) and returns a tuple of
# (industry-formula, benchmark-formula) using the cells from Variables.

def fmt(code):
    return code  # passthrough; openpyxl uses string codes

def build_industry_sheet(title, industry_names):
    ws = wb.create_sheet(title[:31])

    # A2 = industry name string literal so all SUMIFS reference it (so user
    # can rename the group, formulas still work via $B$2-style refs).
    ws['A1'] = 'Industry'
    ws['B1'] = 'Cell ref'
    ws['A1'].font = BOLD
    ws['B1'].font = BOLD
    ws['A2'] = title
    ws['A2'].font = BOLD
    ws['B2'] = 'used by all formulas below ↓'
    ws['B2'].font = Font(italic=True, color='FF6B7280')

    # Metric grid starts at row 4
    HDR_ROW = 4
    ws.cell(row=HDR_ROW, column=1, value='Metric')
    ws.cell(row=HDR_ROW, column=2, value=title)
    ws.cell(row=HDR_ROW, column=3, value=f'{BENCHMARK} (benchmark)')
    ws.cell(row=HDR_ROW, column=4, value='Index vs MFG')
    style_header(ws, row=HDR_ROW)

    # For multi-industry groups (e.g. Google Ads), sum across all underlying
    # industries; for single-industry groups, just one SUMIFS.
    def sumif_group(target_col):
        parts = [
            f'SUMIFS(Variables!{COL[target_col]}2:{COL[target_col]}{LAST_ROW},'
            f' Variables!{COL["industry"]}2:{COL["industry"]}{LAST_ROW},"{ind}")'
            for ind in industry_names
        ]
        return '=' + '+'.join(parts) if len(parts) > 1 else '=' + parts[0]

    def sumif_mfg(target_col):
        return (
            f'=SUMIFS(Variables!{COL[target_col]}2:{COL[target_col]}{LAST_ROW},'
            f' Variables!{COL["industry"]}2:{COL["industry"]}{LAST_ROW},"{BENCHMARK}")'
        )

    # Each row spec: (label, raw_col OR (num_col, denom_col), format_code, section_label_or_None)
    # If raw_col is a string, just SUMIF that. If tuple (num, denom), it's a ratio.
    SPEC = [
        ('EMAIL', None, None, 'section'),
        ('Emails Sent',                'emails_sent',     '#,##0',     None),
        ('Leads Contacted',            'leads_contacted', '#,##0',     None),
        ('Replied (total)',            'replied',         '#,##0',     None),
        ('Reply Rate (per lead)',      ('replied','leads_contacted'),       '0.00%', None),
        ('Bounced',                    'bounced',         '#,##0',     None),
        ('Bounce % (per lead)',        ('bounced','leads_contacted'),       '0.00%', None),
        ('Interested Replies',         'interested',      '#,##0',     None),
        ('Interest Rate (per lead)',   ('interested','leads_contacted'),    '0.0000%', None),

        ('DEMOS', None, None, 'section'),
        ('Demos Booked',               'demos_booked',    '#,##0',     None),
        ('Pending Demos',              'pending',         '#,##0',     None),
        ('Demos / Emails Sent',        ('demos_booked','emails_sent'),      '0.0000%', None),
        ('Demos / Leads Contacted',    ('demos_booked','leads_contacted'),  '0.0000%', None),
        ('Demos / Interested',         ('demos_booked','interested'),       '0.00%', None),

        ('SHOW-UPS', None, None, 'section'),
        ('Total Show-ups',             'showups',         '#,##0',     None),
        ('No-shows',                   'noshow',          '#,##0',     None),
        ('Show-ups / Emails Sent',     ('showups','emails_sent'),           '0.0000%', None),
        ('Show-ups / Leads Contacted', ('showups','leads_contacted'),       '0.0000%', None),
        ('Show-up Rate (Shows / Completed)', ('showups','completed_demos'), '0.00%', None),
        ('Show-ups / Interested',      ('showups','interested'),            '0.00%', None),

        ('CLOSED (ONBOARDINGS)', None, None, 'section'),
        ('Closed Deals',               'closed',          '#,##0',     None),
        ('ARR',                        'arr',             '"$"#,##0',  None),
        ('MRR',                        'mrr',             '"$"#,##0',  None),
        ('Close / Demo',               ('closed','demos_booked'),           '0.00%', None),
        ('Close / Show-up',            ('closed','showups'),                '0.00%', None),
        ('Close / Interested',         ('closed','interested'),             '0.00%', None),
        ('Close / Lead',               ('closed','leads_contacted'),        '0.0000%', None),
    ]

    r = HDR_ROW + 1
    for spec in SPEC:
        label, source, num_fmt, section = spec
        if section == 'section':
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = SECT_FONT
            cell.fill = SECT_FILL
            for col in (2, 3, 4):
                ws.cell(row=r, column=col).fill = SECT_FILL
            r += 1
            continue

        ws.cell(row=r, column=1, value=label)

        if isinstance(source, str):
            # Raw count
            ws.cell(row=r, column=2, value=sumif_group(source))
            ws.cell(row=r, column=3, value=sumif_mfg(source))
        else:
            num_col, denom_col = source
            # Ratio: industry-num / industry-denom (safe-div via IFERROR)
            ws.cell(row=r, column=2,
                    value=f'=IFERROR(({sumif_group(num_col)[1:]})/({sumif_group(denom_col)[1:]}),0)')
            ws.cell(row=r, column=3,
                    value=f'=IFERROR(({sumif_mfg(num_col)[1:]})/({sumif_mfg(denom_col)[1:]}),0)')

        # Index vs MFG = industry / MFG
        ws.cell(row=r, column=4, value=f'=IFERROR(B{r}/C{r},"")')

        # Formats
        ws.cell(row=r, column=2).number_format = num_fmt
        ws.cell(row=r, column=3).number_format = num_fmt
        ws.cell(row=r, column=4).number_format = '0%'

        # Alignment
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).alignment = ALIGN_R
        ws.cell(row=r, column=1).alignment = ALIGN_L

        r += 1

    # Column widths
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 16

    ws.freeze_panes = 'A5'
    return ws

print(f'Building {len(INDUSTRY_GROUPS)} industry pages...', flush=True)
for title, industries in INDUSTRY_GROUPS:
    build_industry_sheet(title, industries)
    print(f'  ✓ {title}', flush=True)

# ── Sheet: Readme ────────────────────────────────────────────────────────
RM = wb.create_sheet('Readme', 0)
RM.append(['Campaign Analysis — Cold Email Insights'])
RM['A1'].font = Font(bold=True, size=14)
RM.append([])
RM.append([f'Source: metrics.json generated {m["generated_at"]}'])
RM.append([f'Benchmark: {BENCHMARK}'])
RM.append(['How it works:'])
RM.append(['  • The Variables sheet is the single source of truth — one row per Sequencer campaign'])
RM.append(['    with raw counts (emails / leads / replied / bounced / interested / demos / shows /'])
RM.append(['    pending / no-shows / completed / closed / ARR / MRR).'])
RM.append(['  • Every industry tab uses SUMIFS to aggregate those raw counts by industry.'])
RM.append(['  • Ratios are computed live as numerator-sum / denominator-sum via IFERROR(...).'])
RM.append(['  • To re-run with fresh data, run:  python3 campaign_analysis.py'])
RM.append([])
RM.append(['Definitions:'])
RM.append(['  Emails Sent          — total deliveries from campaign-stats snapshot'])
RM.append(['  Leads Contacted      — unique leads that received ≥1 email'])
RM.append(['  Replied              — any reply (any tone)'])
RM.append(['  Bounced              — hard-bounce email events'])
RM.append(['  Interested Replies   — Sequencer tag_id=11 (manually flagged Interested)'])
RM.append(['  Demos Booked         — bookings in gist.gtm_inbound_demo_bookings attributed to campaign'])
RM.append(['  Pending Demos        — show_status="P" AND demo_scheduled_date >= today'])
RM.append(['  No-shows             — completed bookings that did not show (N + R + past P)'])
RM.append(['  Show-up Rate         — Shows / (Shows + No-shows)'])
RM.append(['  Closed Deals         — count of demo_bookings flagged as onboarded'])
RM.append(['  ARR / MRR            — sum from Onboarding_Tracker matched on website domain'])
RM.append([])
RM.append(['Index vs MFG = industry value ÷ Manufacturing value (column D on every industry tab)'])
RM.column_dimensions['A'].width = 110

# ── Save ─────────────────────────────────────────────────────────────────
wb.save(OUTFILE)
print(f'\n✓ {OUTFILE}', flush=True)

# Console preview
print()
print(f'{"Industry":<30}{"Campaigns":>12}{"Emails":>14}{"Demos":>8}{"Shows":>8}{"Closed":>8}{"ARR":>12}')
print('-' * 92)
for title, industries in INDUSTRY_GROUPS + [(BENCHMARK, [BENCHMARK])]:
    nc = nem = nd = ns = ncl = na = 0
    for c in m['campaigns']:
        if c['industry'] in industries:
            nc += 1
            nem += c.get('emails_sent') or 0
            cid = c['id']
            nd += demos.get(cid, 0)
            ns += showups.get(cid, 0)
            ncl += closed.get(cid, 0)
            na += arr.get(cid, 0)
    print(f'{title:<30}{nc:>12}{nem:>14,}{nd:>8}{ns:>8}{ncl:>8}${na:>11,.0f}')
