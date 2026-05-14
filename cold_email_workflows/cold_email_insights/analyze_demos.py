"""
analyze_demos.py
────────────────────────────────────────────────────────────────────
Thorough attribution audit. Pulls every gist.gtm_inbound_demo_bookings
record where source ILIKE '%%Gushwork Email%%' (covers the bare label
plus '(Allaine)', '(Paula)', etc.) and tries to map each booking to a
Sequencer campaign → industry using the same progressive matcher the
dashboard ETL uses:

  1. exact email match  → /api/leads/{email}
  2. clean domain match → /api/leads?search={domain}
  3. company-name match → /api/leads?search={company}

Output: data/demo_attribution_audit.xlsx
  • one tab per attributed industry (rows: email, demo date, booking
    date, show status, AE, attribution path, raw source)
  • one tab "Unmapped" with bookings none of the three passes resolved
  • a cover sheet "Summary" with counts

Run from this directory:  python3 analyze_demos.py
"""
import os, sys, psycopg2, psycopg2.extras
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Reuse fetch_campaigns + lookup_lead_attribution from the live ETL so the
# matcher logic is identical to what the dashboard uses.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import etl

DB_URL  = 'postgresql://airbyte_user:airbyte_user_password@gw-rds-analytics.celzx4qnlkfp.us-east-1.rds.amazonaws.com:5432/gw_prod'
OUTFILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'demo_attribution_audit.xlsx')

# ── 1. Pull bookings ──────────────────────────────────────────────────────
print('Fetching bookings with source ILIKE %Gushwork Email%...', flush=True)
conn = psycopg2.connect(DB_URL)
cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
cur.execute("""
    WITH ranked AS (
        SELECT prospect_email, prospect_company, prospect_website,
               demo_scheduled_date::text       AS demo_date,
               created_at::date::text          AS booking_date,
               show_status, ae_name, source,
               ROW_NUMBER() OVER (
                 PARTITION BY LOWER(prospect_email)
                 ORDER BY demo_scheduled_date DESC NULLS LAST,
                          CASE show_status WHEN 'Y' THEN 1 WHEN 'P' THEN 2
                                           WHEN 'N' THEN 3 WHEN 'R' THEN 4
                                           ELSE 5 END,
                          created_at DESC
               ) AS rn
        FROM gist.gtm_inbound_demo_bookings
        WHERE is_latest = true
          AND prospect_email IS NOT NULL
          AND source ILIKE %s
    )
    SELECT LOWER(prospect_email) AS email,
           prospect_company       AS company,
           prospect_website       AS website,
           demo_date, booking_date, show_status, ae_name, source
    FROM ranked
    WHERE rn = 1
    ORDER BY demo_date DESC NULLS LAST
""", ('%Gushwork Email%',))
bookings = cur.fetchall()
conn.close()
print(f'  {len(bookings)} unique bookings', flush=True)

# ── 2. Load Sequencer campaigns ───────────────────────────────────────────
print('Loading Sequencer campaigns...', flush=True)
campaigns       = etl.fetch_campaigns()
campaigns_by_id = {c['id']: c for c in campaigns}
print(f'  {len(campaigns)} campaigns', flush=True)

# ── 3. Attribute in parallel ──────────────────────────────────────────────
print(f'Attributing {len(bookings)} bookings (8 workers)...', flush=True)

def attribute(b):
    ind, cid, src = etl.lookup_lead_attribution(b['email'], b['company'], campaigns_by_id)
    camp = campaigns_by_id.get(cid) if cid else None
    return {
        **b,
        'attr_industry':    ind,
        'attr_campaign_id': cid,
        'attr_campaign':    camp['name'] if camp else '',
        'attr_path':        src,
    }

results = []
with ThreadPoolExecutor(max_workers=8) as ex:
    futs = {ex.submit(attribute, b): b for b in bookings}
    for i, f in enumerate(as_completed(futs), 1):
        results.append(f.result())
        if i % 25 == 0 or i == len(bookings):
            print(f'  [{i}/{len(bookings)}]', flush=True)

# ── 4. Bucket by industry ─────────────────────────────────────────────────
by_ind = defaultdict(list)
for r in results:
    by_ind[r['attr_industry'] or 'Unmapped'].append(r)

# ── 5. Write Excel ────────────────────────────────────────────────────────
print(f'\nWriting {OUTFILE}', flush=True)

wb = Workbook()
wb.remove(wb.active)

HEADER_FONT = Font(bold=True, color='FFFFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='FF0070FF')
HEADER_ALIGN = Alignment(horizontal='left', vertical='center')
THIN = Side(style='thin', color='FFD0D5DD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws):
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
        c.border = BORDER

def autosize(ws, max_w=50):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_w)

# 5a. Summary sheet (first tab)
summary = wb.create_sheet('Summary')
summary.append(['Industry', 'Bookings', 'via email', 'via domain', 'via company', 'Unmapped'])
style_header_row(summary)

industries_sorted = sorted(
    [k for k in by_ind.keys() if k != 'Unmapped'],
    key=lambda k: -len(by_ind[k]),
)
if 'Unmapped' in by_ind:
    industries_sorted.append('Unmapped')

def path_counts(rows):
    c = defaultdict(int)
    for r in rows:
        c[r['attr_path']] += 1
    return c

for ind in industries_sorted:
    rows = by_ind[ind]
    pc = path_counts(rows)
    summary.append([
        ind, len(rows),
        pc.get('email', 0), pc.get('domain', 0),
        pc.get('company', 0), pc.get('none', 0),
    ])
summary.append([])
summary.append(['TOTAL', len(results),
                sum(1 for r in results if r['attr_path']=='email'),
                sum(1 for r in results if r['attr_path']=='domain'),
                sum(1 for r in results if r['attr_path']=='company'),
                sum(1 for r in results if r['attr_path']=='none')])
for c in summary[summary.max_row]:
    c.font = Font(bold=True)
autosize(summary)

# 5b. Per-industry tabs
COLS = ['Email', 'Demo Date', 'Booking Date', 'Show Status', 'AE Name',
        'Company', 'Website', 'Attribution Path', 'Campaign', 'Source (DB)']

def sheet_name(s):
    # Excel sheet names: ≤31 chars, no : \ / ? * [ ]
    name = s.replace('/', '-').replace(':', '-')
    for ch in ['\\', '?', '*', '[', ']']:
        name = name.replace(ch, '')
    return name[:31] or 'Sheet'

for ind in industries_sorted:
    ws = wb.create_sheet(sheet_name(ind))
    ws.append(COLS)
    style_header_row(ws)
    for r in sorted(by_ind[ind], key=lambda x: (x.get('demo_date') or '', x['email']), reverse=True):
        ws.append([
            r['email'],
            r.get('demo_date') or '',
            r.get('booking_date') or '',
            r.get('show_status') or '',
            r.get('ae_name') or '',
            r.get('company') or '',
            r.get('website') or '',
            r.get('attr_path') or '',
            r.get('attr_campaign') or '',
            r.get('source') or '',
        ])
    autosize(ws)

# Freeze header on every sheet
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'

wb.save(OUTFILE)

# ── 6. Console summary ────────────────────────────────────────────────────
print(f'\n✓ {OUTFILE}\n')
print(f'{"Industry":<35}{"Total":>7}{"email":>8}{"domain":>8}{"company":>9}{"unmapped":>10}')
print('-' * 78)
for ind in industries_sorted:
    rows = by_ind[ind]
    pc = path_counts(rows)
    print(f'{ind:<35}{len(rows):>7}{pc.get("email",0):>8}{pc.get("domain",0):>8}{pc.get("company",0):>9}{pc.get("none",0):>10}')
print('-' * 78)
print(f'{"TOTAL":<35}{len(results):>7}'
      f'{sum(1 for r in results if r["attr_path"]=="email"):>8}'
      f'{sum(1 for r in results if r["attr_path"]=="domain"):>8}'
      f'{sum(1 for r in results if r["attr_path"]=="company"):>9}'
      f'{sum(1 for r in results if r["attr_path"]=="none"):>10}')
